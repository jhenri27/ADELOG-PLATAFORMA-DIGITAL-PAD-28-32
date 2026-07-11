import openpyxl
import json

excel_path = r"C:\wamp64\www\PLATAFORMA DIGITAL-PAD-28-32\DATOS ELECTORALES\Resultado Diputados por colegios Elecciones Presidenciales y Congresuales 2024 (7) (6) (1).xlsx"
json_output = r"C:\wamp64\www\PLATAFORMA DIGITAL-PAD-28-32\scratch\colegios_mapping.json"

print("Loading Workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

target_sheets = [
    "VOTOS DE SAN ANT. GUERRA",
    "VOTOS DE SAN LUIS",
    "VOTOS DE LA CALETA",
    "VOTOS DE BOCA CHICA ",
    "REGION 3 ",
    "REGION 3A ",
    "REGION 3B ",
    "REGION 3C ",
    "REGION 3D ",
    "REGION 3E ",
    "REGION 3F"
]

colegios_data = []

for name in wb.sheetnames:
    if name not in target_sheets:
        continue
    
    sheet = wb[name]
    print(f"Reading sheet: {name}")
    
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 2:
            continue
            
        col_number = row[1]
        recinto = row[2]
        region = row[3]
        zona = row[4]
        
        if not col_number or str(col_number).strip() == "" or "total" in str(col_number).lower():
            continue
            
        v_prm = 0
        v_dip = 0
        try:
            v_prm = int(row[6]) if row[6] is not None else 0
            v_dip = int(row[7]) if row[7] is not None else 0
        except:
            pass
            
        colegios_data.append({
            "colegio": str(col_number).strip().upper(),
            "recinto": str(recinto).strip() if recinto else "N/A",
            "region": str(region).strip() if region else "N/A",
            "zona": str(zona).strip() if zona else "N/A",
            "votos_prm": v_prm,
            "votos_diputada": v_dip
        })

with open(json_output, "w", encoding="utf-8") as f:
    json.dump(colegios_data, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(colegios_data)} records to: {json_output}")
