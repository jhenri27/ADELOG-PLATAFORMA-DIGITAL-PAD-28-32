import openpyxl
import mysql.connector

excel_path = r"C:\wamp64\www\PLATAFORMA DIGITAL-PAD-28-32\DATOS ELECTORALES\Resultado Diputados por colegios Elecciones Presidenciales y Congresuales 2024 (7) (6) (1).xlsx"

print("Loading Workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Connect to MySQL database
print("Connecting to database...")
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="pad_electoral_2832"
)
cursor = conn.cursor()

# We only care about sheets that list individual colegios
target_sheets = [
    "VOTOS DE SAN ANT. GUERRA",
    "VOTOS DE SAN LUIS",
    "VOTOS DE LA CALETA",
    "VOTOS DE BOCA CHICA ",
    "REGION 3 ",
    "REGION 3A ",
    "REGION 3B ",
    "REGION 3C ",
    "REGION 3D ",
    "REGION 3E ",
    "REGION 3F"
]

total_inserted = 0

for name in wb.sheetnames:
    if name not in target_sheets:
        continue
    
    sheet = wb[name]
    print(f"Processing sheet: {name}")
    
    # Iterate rows
    # Col index mapping for these sheets:
    # 0: NUM., 1: COLEGIO, 2: RECINTO ELECTORAL, 3: REGION, 4: ZONA, 5: CANT.ELECTORES, 6: VOTOS PRM, 7: CANTIDAD VOTOS DIPUTADAS
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 2:
            continue # Skip header/title rows
            
        col_number = row[1]
        recinto = row[2]
        region = row[3]
        zona = row[4]
        
        # Check if this row is valid and not a summary/total row
        if not col_number or str(col_number).strip() == "" or "total" in str(col_number).lower():
            continue
            
        # Clean college number
        colegio_str = str(col_number).strip().upper()
        recinto_str = str(recinto).strip() if recinto else "N/A"
        region_str = str(region).strip() if region else "N/A"
        zona_str = str(zona).strip() if zona else "N/A"
        
        votos_prm = 0
        votos_dip = 0
        try:
            votos_prm = int(row[6]) if row[6] is not None else 0
            votos_dip = int(row[7]) if row[7] is not None else 0
        except:
            pass
            
        # Insert or update
        query = """
            INSERT INTO colegios_estructural (colegio, recinto, region, zona, votos_prm, votos_diputada)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                recinto = VALUES(recinto),
                region = VALUES(region),
                zona = VALUES(zona),
                votos_prm = VALUES(votos_prm),
                votos_diputada = VALUES(votos_diputada)
        """
        cursor.execute(query, (colegio_str, recinto_str, region_str, zona_str, votos_prm, votos_dip))
        total_inserted += 1

conn.commit()
cursor.close()
conn.close()

print(f"ETL Complete: Loaded {total_inserted} Colegios into colegios_estructural database table.")
