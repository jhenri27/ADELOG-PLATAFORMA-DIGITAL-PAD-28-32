import openpyxl

excel_path = r"C:\wamp64\www\PLATAFORMA DIGITAL-PAD-28-32\DATOS ELECTORALES\Resultado Diputados por colegios Elecciones Presidenciales y Congresuales 2024 (7) (6) (1).xlsx"
print("Loading Workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Worksheets found:")
for name in wb.sheetnames:
    print(f"- {name}")
    sheet = wb[name]
    print("  First 5 rows:")
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 5:
            print(f"    Row {r_idx+1}: {row}")
        else:
            break
