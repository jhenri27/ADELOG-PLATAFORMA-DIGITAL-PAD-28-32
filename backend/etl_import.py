import openpyxl
import json
import os

excel_path = r"C:\wamp64\www\PLATAFORMA DIGITAL-PAD-28-32\DATOS ELECTORALES\Resultado Diputados por colegios Elecciones Presidenciales y Congresuales 2024 (7) (6) (1).xlsx"
json_output_path = r"C:\wamp64\www\PLATAFORMA DIGITAL-PAD-28-32\backend\historical_data.json"

print("ETL STARTING...")
if not os.path.exists(excel_path):
    print(f"ERROR: Excel file not found at: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Locate the summary sheet
sheet_name = None
for name in wb.sheetnames:
    if "PORSENTAJE" in name or "ORDEN" in name or "VOTACION" in name:
        sheet_name = name
        break

if not sheet_name:
    print("ERROR: Could not find sheet with percentages and order of voting.")
    exit(1)

sheet = wb[sheet_name]
print(f"Reading sheet: {sheet_name}")

historical_data = []

# Iterate rows, headers are: #, REGION, VOTOS PRM, VOTOS DIPUTADA, PORCIENTO
for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if r_idx == 0:
        continue # Skip header row
        
    num = row[0]
    region = row[1]
    votos_prm = row[2]
    votos_diputada = row[3]
    porciento = row[4]
    
    # If region is empty or this is the total row, break
    if not region or str(region).strip() == "" or "total" in str(region).lower():
        # Stop if we hit a row without region name or the summary total row
        if r_idx > 1 and (region is None or "total" in str(region).lower()):
            break
        continue
        
    # Convert and clean values
    try:
        v_prm = int(votos_prm) if votos_prm is not None else 0
        v_dip = int(votos_diputada) if votos_diputada is not None else 0
        pct = float(porciento) if porciento is not None else 0.0
        
        region_clean = str(region).strip().upper()
        
        historical_data.append({
            "region": region_clean,
            "votos_prm": v_prm,
            "votos_diputada": v_dip,
            "porciento": pct
        })
        print(f"  Parsed Region: {region_clean} | Votos PRM: {v_prm} | Votos Diputada: {v_dip} | Pct: {pct:.4f}")
    except Exception as e:
        print(f"  Warning: Row {r_idx+1} could not be parsed: {row}. Error: {e}")

# Write to JSON
with open(json_output_path, "w", encoding="utf-8") as f:
    json.dump(historical_data, f, ensure_ascii=False, indent=2)

print(f"ETL COMPLETE. JSON written with {len(historical_data)} records to: {json_output_path}")
