import openpyxl
import csv
import json
import sys
import os
import argparse

def read_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        headers = []
        for row in sheet.iter_rows(max_row=1, values_only=True):
            headers = [str(cell).strip() if cell is not None else f"Columna_{i+1}" for i, cell in enumerate(row)]
            break
        
        # Read up to 3 sample rows
        samples = []
        row_count = 0
        for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
            samples.append([str(cell).strip() if cell is not None else "" for cell in row])
            row_count += 1
            
        return {
            "exito": True,
            "formato": "Excel",
            "hojas": wb.sheetnames,
            "columnas": headers,
            "muestras": samples
        }
    except Exception as e:
        return {"exito": False, "mensaje": f"Error al leer Excel: {str(e)}"}

def read_csv(filepath):
    try:
        # Detect delimiter
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            sample = f.read(2048)
            dialect = csv.Sniffer().sniff(sample)
            delimiter = dialect.delimiter
        
        headers = []
        samples = []
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [str(cell).strip() if cell else f"Columna_{idx+1}" for idx, cell in enumerate(row)]
                elif i <= 3:
                    samples.append([str(cell).strip() if cell else "" for cell in row])
                else:
                    break
                    
        return {
            "exito": True,
            "formato": "CSV",
            "columnas": headers,
            "muestras": samples
        }
    except Exception as e:
        return {"exito": False, "mensaje": f"Error al leer CSV: {str(e)}"}

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--file', required=True)
    args = parser.parse_args()
    
    filepath = args.file
    if not os.path.exists(filepath):
        print(json.dumps({"exito": False, "mensaje": "Archivo no encontrado."}))
        sys.exit(1)
        
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xlsx':
        res = read_excel(filepath)
    elif ext in ['.csv', '.txt']:
        res = read_csv(filepath)
    else:
        res = {"exito": False, "mensaje": "Formato de archivo no soportado."}
        
    print(json.dumps(res, ensure_ascii=False))
